"""Bulk content import service. Supports CSV, XLSX, and JSON."""
import csv
import io
import json
from datetime import datetime, timezone

from fastapi import HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.content import ExamType, SkillTag, Subject, Topic
from app.models.curriculum import CurriculumOutcome, QuestionOutcomeMapping
from app.models.import_job import ImportJob, ImportJobStatus
from app.models.question import (
    ContentOwnershipType,
    DifficultyLevel,
    Question,
    QuestionStatus,
    QuestionType,
    QuestionVersion,
    SourceType,
)

_REQUIRED_FIELDS = {"question_text", "answer", "subject", "exam_type"}
_VALID_DIFFICULTIES = {d.value for d in DifficultyLevel}
_VALID_SOURCES = {"manual", "ocr", "ai", "imported"}


# ── Parsing ──────────────────────────────────────────────────────────────────


async def parse_file(file: UploadFile) -> list[dict]:
    """Parse uploaded file into list of row dicts. Raises 422 on invalid format."""
    content = await file.read()
    filename = (file.filename or "").lower()

    if filename.endswith(".csv"):
        return _parse_csv(content)
    elif filename.endswith(".xlsx"):
        return _parse_xlsx(content)
    elif filename.endswith(".json"):
        return _parse_json(content)
    else:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Unsupported file format. Use .csv, .xlsx, or .json",
        )


def _parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [normalize_row(r) for r in reader]


def _parse_xlsx(content: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel support not installed (missing openpyxl)")

    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    if not ws:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = [str(h).strip().lower().replace(" ", "_") if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(c is None for c in row):
            continue
        values = [str(c) if c is not None else "" for c in row]
        if len(values) < len(headers):
            values += [""] * (len(headers) - len(values))
        result.append(normalize_row(dict(zip(headers, values))))
    return result


def _parse_json(content: bytes) -> list[dict]:
    try:
        data = json.loads(content.decode("utf-8-sig"))
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=422, detail=f"Invalid JSON: {e}")

    if isinstance(data, dict) and "questions" in data:
        rows = data["questions"]
    elif isinstance(data, list):
        rows = data
    else:
        raise HTTPException(status_code=422, detail="JSON must be an array or { questions: [...] }")

    if not isinstance(rows, list):
        raise HTTPException(status_code=422, detail="JSON questions must be an array")
    return [normalize_row(r) for r in rows]


def normalize_row(row: dict) -> dict:
    """Normalize keys: lower-case, replace spaces/hyphens with underscore."""
    out = {}
    for k, v in row.items():
        key = str(k).strip().lower().replace(" ", "_").replace("-", "_")
        out[key] = str(v).strip() if v else ""
    return out


# ── Validation ───────────────────────────────────────────────────────────────


async def validate_rows(
    rows: list[dict], db: AsyncSession, skip_duplicates: bool = True
) -> dict:
    """Validate all rows. Returns { valid, invalid, duplicates, preview_items }."""
    # Preload lookup caches
    subjects = {s.code.lower(): s for s in (await db.execute(select(Subject))).scalars().all()}
    subjects.update({s.name.lower(): s for s in subjects.values()})

    exam_types = {et.code.lower(): et for et in (await db.execute(select(ExamType))).scalars().all()}
    exam_types.update({et.name.lower(): et for et in exam_types.values()})

    topics = {(t.subject_id, t.name.lower()): t for t in (await db.execute(select(Topic))).scalars().all()}
    skills = {(s.subject_id, s.name.lower()): s for s in (await db.execute(select(SkillTag))).scalars().all()}
    outcomes = {o.code: o for o in (await db.execute(select(CurriculumOutcome))).scalars().all()}

    # Existing question fingerprints for duplicate detection
    existing = await db.execute(
        select(Question.subject_id, Question.exam_type_id, QuestionVersion.stem)
        .join(QuestionVersion, QuestionVersion.question_id == Question.id)
        .where(QuestionVersion.version_number == 1)
    )
    existing_fingerprints = {(row[0], row[1], row[2].strip().lower()) for row in existing.fetchall()}

    valid = []
    invalid = []
    duplicates = []
    seen = set()

    for i, row in enumerate(rows):
        errors = []
        # Required fields
        for field in _REQUIRED_FIELDS:
            if not row.get(field):
                errors.append(f"Missing required field: {field}")

        # Subject lookup
        subject_key = row.get("subject", "").lower()
        subject = subjects.get(subject_key)
        if not subject:
            errors.append(f"Subject not found: '{row.get('subject')}'")

        # Exam type lookup
        et_key = row.get("exam_type", "").lower()
        exam_type = exam_types.get(et_key)
        if not exam_type:
            errors.append(f"Exam type not found: '{row.get('exam_type')}'")

        # Difficulty
        difficulty = row.get("difficulty", "medium").lower()
        if difficulty and difficulty not in _VALID_DIFFICULTIES:
            errors.append(f"Invalid difficulty: '{difficulty}'")

        # Source
        source = row.get("source_type", "imported").lower()
        if source not in _VALID_SOURCES:
            errors.append(f"Invalid source_type: '{source}'")

        if errors:
            invalid.append({"row": i + 1, "errors": errors, "data": row})
            continue

        # Duplicate check
        if subject and exam_type:
            fingerprint = (subject.id, exam_type.id, row.get("question_text", "").strip().lower())
            if fingerprint in existing_fingerprints or fingerprint in seen:
                duplicates.append({"row": i + 1, "question_text": row.get("question_text", "")})
                if skip_duplicates:
                    continue
        if subject and exam_type:
            seen.add((subject.id, exam_type.id, row.get("question_text", "").strip().lower()))

        # Topic lookup
        topic = None
        topic_name = row.get("topic", "")
        if topic_name and subject:
            topic = topics.get((subject.id, topic_name.lower()))

        # Skill lookup
        skill = None
        skill_name = row.get("skill", "")
        if skill_name and subject:
            skill = skills.get((subject.id, skill_name.lower()))

        # Outcome lookup
        outcome_code = row.get("curriculum_outcome", "")
        outcome = outcomes.get(outcome_code) if outcome_code else None

        valid.append({
            "row": i + 1,
            "stem": row.get("question_text", ""),
            "correct_answer": row.get("answer", ""),
            "difficulty": difficulty or "medium",
            "subject_id": subject.id if subject else "",
            "exam_type_id": exam_type.id if exam_type else "",
            "topic_id": topic.id if topic else None,
            "topic_name": topic.name if topic else "",
            "skill_id": skill.id if skill else None,
            "skill_name": skill.name if skill else "",
            "outcome_code": outcome_code,
            "outcome_id": outcome.id if outcome else None,
            "explanation": row.get("explanation", ""),
            "source_type": source or "imported",
            "year_level": 5,
            "options_json": _build_mcq_options(row.get("answer", "")),
        })

    return {
        "total_rows": len(rows),
        "valid_count": len(valid),
        "invalid_count": len(invalid),
        "duplicate_count": len(duplicates),
        "valid": valid,
        "invalid": invalid,
        "duplicates": duplicates,
    }


def _build_mcq_options(answer: str) -> list[dict]:
    """Generate 4 options from answer. Default distractor generation."""
    correct = answer.strip()
    if not correct:
        return []
    return [
        {"label": "A", "text": correct, "is_correct": True, "explanation": ""},
        {"label": "B", "text": f"Not {correct}", "is_correct": False, "explanation": ""},
        {"label": "C", "text": f"Neither {correct}", "is_correct": False, "explanation": ""},
        {"label": "D", "text": "None of the above", "is_correct": False, "explanation": ""},
    ]


# ── Import Execution ─────────────────────────────────────────────────────────


async def execute_import(
    valid_rows: list[dict],
    admin_id: str,
    filename: str,
    file_format: str,
    db: AsyncSession,
) -> ImportJob:
    """Create questions from validated rows. All imported as draft."""
    job = ImportJob(
        filename=filename,
        format=file_format,
        uploaded_by=admin_id,
        status=ImportJobStatus.processing,
        started_at=datetime.now(tz=timezone.utc),
    )
    db.add(job)
    await db.commit()

    imported = 0
    mappings = 0
    skipped = 0
    failed = 0

    for row in valid_rows:
        try:
            q = Question(
                subject_id=row["subject_id"],
                exam_type_id=row["exam_type_id"],
                year_level=row.get("year_level", 5),
                topic_id=row.get("topic_id"),
                difficulty=row.get("difficulty", "medium"),
                question_type=QuestionType.mcq,
                status=QuestionStatus.draft,
                source_type=SourceType(row.get("source_type", "imported")),
                content_ownership=ContentOwnershipType.original,
                created_by_admin_id=admin_id,
                current_version_id=None,
            )
            db.add(q)
            await db.flush()

            v = QuestionVersion(
                question_id=q.id,
                version_number=1,
                stem=row["stem"],
                correct_answer=row["correct_answer"],
                full_explanation=row.get("explanation", ""),
                marks=1,
                options_json=row.get("options_json"),
                created_by_admin_id=admin_id,
                created_at=datetime.now(tz=timezone.utc),
            )
            db.add(v)
            await db.flush()
            q.current_version_id = v.id
            imported += 1

            # Curriculum outcome mapping
            if row.get("outcome_id"):
                mapping = QuestionOutcomeMapping(
                    question_id=q.id,
                    outcome_id=row["outcome_id"],
                    weight=1.0,
                )
                db.add(mapping)
                mappings += 1

        except Exception:
            failed += 1
            continue

    job.status = ImportJobStatus.completed
    job.imported_count = imported
    job.skipped_count = skipped
    job.failed_count = failed
    job.duplicate_count = 0  # already filtered before execution
    job.mapping_count = mappings
    job.completed_at = datetime.now(tz=timezone.utc)

    await db.commit()
    await db.refresh(job)
    return job


async def list_import_jobs(db: AsyncSession) -> list[ImportJob]:
    result = await db.execute(
        select(ImportJob).order_by(ImportJob.created_at.desc()).limit(50)
    )
    return list(result.scalars().all())


async def get_import_job(job_id: str, db: AsyncSession) -> ImportJob:
    result = await db.execute(select(ImportJob).where(ImportJob.id == job_id))
    job = result.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail="Import job not found")
    return job
